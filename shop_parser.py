from bs4 import BeautifulSoup
import urllib.request
import openpyxl
import requests

def save_img(img, name, package):
    p = requests.get('http://martika-home.ru' + img)
    out = open(r'C:/Users/sergey/Documents/' + package + '/' + name.replace('\t', '') + '.jpg', 'wb')
    out.write(p.content)
    out.close()

def get_html(url):
    response = urllib.request.urlopen(url)
    return response.read()

def parse(html, products, package):
    soup = BeautifulSoup(html, 'html.parser')

    table = soup.find('table', class_='product-item')

    links = table.find_all('tr', class_='product-item-link')
    img_links = table.find_all('tr', class_='product-item-img')
    imgs = []
    for link in img_links:
        img_link = link.find_all('td')
        for src in img_link:
            img = src.img['src']
            imgs.append(img)
    titles_text = []
    for link in links:
        titles = link.find_all('a', class_='black')
        for title in titles:
            titles_text.append(title.text)
    i = 0
    for text in titles_text:
        word_array = text.split(' ')
        articul = word_array[-1]
        product = {
            'title': ' '.join(word_array[:-1]),
            'articul': articul,
            'web_src' : 'http://martika-home.ru' + imgs[i],
            'local_src': 'C:/Users/sergey/Documents/' + package + '/' + articul.replace('\t', '') + '.jpg'
        }
        save_img(imgs[i], articul, package)
        products.append(product)
        i += 1


def parse_num_list(html):
    soup = BeautifulSoup(html, 'html.parser')
    a = soup.find_all('a')
    return a[-7].text

def main():
    first_page = 1
    hos_bit_products = []
    kitchen_products = []
    garden_products = []
    winter = []

    num_pages_hos_bit = int(parse_num_list(get_html('http://martika-home.ru/catalog/list.php?SECTION_ID=152')))
    while (first_page <= num_pages_hos_bit):
        parse(get_html('http://martika-home.ru/catalog/list.php?SECTION_ID=152&PAGEN_1=' + str(first_page)), hos_bit_products, 'HosTovars')
        first_page += 1

    first_page = 1
    num_pages_kitchen = int(parse_num_list(get_html('http://martika-home.ru/catalog/list.php?SECTION_ID=151')))
    while (first_page <= num_pages_kitchen):
        parse(get_html('http://martika-home.ru/catalog/list.php?SECTION_ID=151&PAGEN_1=' + str(first_page)), kitchen_products, 'KitchenTovars')
        first_page += 1

    first_page = 1
    num_pages_garden = int(parse_num_list(get_html('http://martika-home.ru/catalog/list.php?SECTION_ID=150')))
    while (first_page <= num_pages_garden):
        parse(get_html('http://martika-home.ru/catalog/list.php?SECTION_ID=150&PAGEN_1=' + str(first_page)), garden_products, 'GardenTovars')
        first_page += 1

    parse(get_html('http://martika-home.ru/catalog/list.php?SECTION_ID=149'), winter, 'WinterTovars')
    save_in_excel(hos_bit_products, 'HosTovars')
    save_in_excel(kitchen_products, 'KitchenTovars')
    save_in_excel(garden_products, 'GardenTovars')
    save_in_excel(winter, 'WinterTovars')



def save_in_excel(obj, list_name):
    wb = openpyxl.load_workbook(filename = 'C:/Users/sergey/Documents/Products.xlsx')
    sheet = wb[list_name]
    i = 2
    for dict in obj:
        for key, value in dict.items():
            if key == 'title':
                sheet.cell(row=i, column=1).value = value
            if key == 'articul':
                sheet.cell(row=i, column=2).value = value
            if key == 'web_src':
                sheet.cell(row=i, column=3).value = value
            if key == 'local_src':
                sheet.cell(row=i, column=4).value = value
        i+=1
    wb.save(r'C:/Users/sergey/Documents/Products.xlsx')



if __name__ == '__main__':
    main()




